from datetime import datetime
from enum import Enum
from io import BytesIO

from fastapi import Depends, HTTPException
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from sqlalchemy import select, update, text
from sqlalchemy.exc import NoResultFound
from sqlalchemy.ext.asyncio import AsyncSession

from ..database.database import get_db
from ..database.models import Order, OrderState


class OrdersRepository:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def create_order(self, order: Order):
        try:
            self.db.add(order)
            await self.db.commit()
            await self.db.refresh(order)
            return order
        except Exception as e:
            await self.db.rollback()
            raise e

    async def get_orders(self):
        try:
            result = await self.db.execute(select(Order))
            orders = result.scalars().all()
            return orders
        except Exception as e:
            raise HTTPException(status_code=500, detail="Error fetching orders")

    async def get_order_by_id(self, order_id: int):
        try:
            result = await self.db.execute(select(Order).filter(Order.id == order_id))
            order = result.scalar_one_or_none()  # Fetch one or return None if not found
            if not order:
                raise NoResultFound(f"Order with id {order_id} not found")
            return order
        except NoResultFound as e:
            raise HTTPException(status_code=404, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=500, detail="Error fetching orders by id")

    async def delete_order(self, order_id: int):
        try:
            order = await self.get_order_by_id(order_id)  # Check if order exists
            await self.db.delete(order)
            await self.db.commit()
            return {"detail": f"Order with id {order_id} deleted successfully"}
        except NoResultFound:
            raise HTTPException(status_code=404, detail=f"Order with id {order_id} not found")
        except Exception as e:
            await self.db.rollback()
            raise HTTPException(status_code=500, detail="Error deleting order")

    async def delete_all_orders(self):
        try:
            orders = await self.db.execute('SELECT * FROM orders')
            orders_list = orders.fetchall()

            if not orders_list:
                raise HTTPException(status_code=404, detail="No orders found to delete")


            await self.db.execute('DELETE FROM orders')
            await self.db.commit()
            return {"detail": "All orders deleted successfully"}
        except Exception as e:
            await self.db.rollback()
            raise HTTPException(status_code=500, detail="Error deleting orders")

    async def change_state(self, order_id: int, new_state: OrderState):
        try:
            stmt = (
                update(Order)
                .where(Order.id == order_id)
                .values(state=new_state)
                .execution_options(synchronize_session="fetch")
            )
            result = await self.db.execute(stmt)
            await self.db.commit()

            if result.rowcount == 0:
                raise NoResultFound(f"Order with id {order_id} not found")

            return {"detail": f"Order with id {order_id} updated to state {new_state}"}
        except NoResultFound as e:
            raise HTTPException(status_code=404, detail=str(e))
        except Exception as e:
            await self.db.rollback()
            raise HTTPException(status_code=500, detail="Error changing state of order")

    async def get_excel_data(self, export_path):
        result = await self.db.execute(select(Order))
        orders = result.scalars().all()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Orders"

        # Exclude the 'id' column from headers
        headers = [col.name for col in Order.__table__.columns if col.name != 'id']
        sheet.append(headers)

        for order in orders:
            row_data = [
                getattr(order, col.name).value if isinstance(getattr(order, col.name), Enum)
                else getattr(order, col.name)
                for col in Order.__table__.columns if col.name != 'id'  # Exclude 'id' in row data as well
            ]
            sheet.append(row_data)

        workbook.save(export_path)

    async def import_orders_from_excel(self, file: BytesIO):
        try:
            workbook = load_workbook(file)
            sheet = workbook.active

            # Получаем заголовки
            headers = [cell.value for cell in sheet[1]]

            # Exclude 'id' column from expected headers
            expected_headers = ["name", "number", "date", "client_name", "state", "order_amount"]

            # Check that the headers match the expected ones (without 'id')
            if sorted(headers) != sorted(expected_headers):
                raise HTTPException(status_code=400, detail="Invalid Excel file format")

            # Iterate over rows (starting from the second row since the first row is the header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Create a dictionary excluding the 'id' column from the data
                order_data = dict(zip(headers, row))

                # Ensure the 'date' field is properly formatted
                if isinstance(order_data["date"], datetime):
                    order_data["date"] = order_data["date"].strftime("%d-%m-%Y")

                # Convert the order status string to the appropriate Enum value (if necessary)
                state = OrderState(order_data["state"])

                # Create a new Order object (exclude the 'id' column here)
                order = Order(
                    name=order_data["name"],
                    number=order_data["number"],
                    date=order_data["date"],
                    client_name=order_data["client_name"],
                    state=state,
                    order_amount=order_data["order_amount"],
                )

                # Add the order to the session for database insertion
                self.db.add(order)

            # Commit the changes to the database
            await self.db.commit()

            return {"detail": "Orders successfully imported"}

        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

async def get_orders_repository(db: AsyncSession = Depends(get_db)):
    return OrdersRepository(db)